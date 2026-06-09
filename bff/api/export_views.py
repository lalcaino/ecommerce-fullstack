import io
import logging
from datetime import datetime
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from .gateway import MicroserviceGateway

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─── Helpers de estilo ────────────────────────────────────────────────────────

VERDE        = '408A71'
VERDE_CLARO  = 'E8F5F0'
GRIS         = 'F3F4F6'
BLANCO       = 'FFFFFF'
ROJO         = 'EF4444'
AMARILLO     = 'F59E0B'

def _header_fill(color=VERDE):
    return PatternFill('solid', start_color=color, end_color=color)

def _header_font(bold=True, color=BLANCO):
    return Font(name='Arial', bold=bold, color=color, size=10)

def _data_font(bold=False):
    return Font(name='Arial', bold=bold, size=10)

def _border():
    thin = Side(style='thin', color='D1D5DB')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _apply_header_row(ws, headers, row=1, fill_color=VERDE):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font      = _header_font(color=BLANCO if fill_color == VERDE else '1F2937')
        cell.fill      = _header_fill(fill_color)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[row].height = 22

def _apply_data_row(ws, values, row, alt=False):
    fill = PatternFill('solid', start_color=VERDE_CLARO if alt else BLANCO,
                       end_color=VERDE_CLARO if alt else BLANCO)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _data_font()
        cell.fill      = fill
        cell.alignment = _left()
        cell.border    = _border()

def _set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def _titulo(ws, texto, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=texto)
    cell.font      = Font(name='Arial', bold=True, size=13, color=BLANCO)
    cell.fill      = _header_fill(VERDE)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28

def _subtitulo(ws, texto, n_cols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=texto)
    cell.font      = Font(name='Arial', size=9, color='6B7280')
    cell.alignment = _center()
    ws.row_dimensions[row].height = 16


# ─── Hojas ────────────────────────────────────────────────────────────────────

def _hoja_resumen(wb, pedidos, productos, bodegas, tiendas, envios, empresa_rut):
    ws = wb.active
    ws.title = '📊 Resumen'

    fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
    _titulo(ws, 'SmartLogix — Reporte de Negocio', 4)
    _subtitulo(ws, f'Generado el {fecha} | Empresa: {empresa_rut or "Todas"}', 4)

    # Métricas
    headers = ['Métrica', 'Valor', 'Detalle', 'Estado']
    _apply_header_row(ws, headers, row=4, fill_color='1F2937')

    total_pedidos    = len(pedidos)
    pendientes       = sum(1 for p in pedidos if p.get('estado') == 'PENDIENTE')
    procesando       = sum(1 for p in pedidos if p.get('estado') == 'PROCESANDO')
    enviados         = sum(1 for p in pedidos if p.get('estado') == 'ENVIADO')
    entregados       = sum(1 for p in pedidos if p.get('estado') == 'ENTREGADO')
    cancelados       = sum(1 for p in pedidos if p.get('estado') == 'CANCELADO')
    total_ventas     = sum(float(p.get('total', 0)) for p in pedidos if p.get('estado') != 'CANCELADO')
    total_productos  = len(productos)
    bajo_stock       = sum(1 for p in productos if p.get('stock', 0) <= p.get('stock_minimo', 5))
    total_bodegas    = len(bodegas)
    total_tiendas    = len(tiendas)
    envios_en_ruta   = sum(1 for e in envios if e.get('estado') == 'EN_RUTA')

    metricas = [
        ('Total Pedidos',        total_pedidos,   f'{entregados} entregados',          '✅' if total_pedidos > 0 else '—'),
        ('Pedidos Pendientes',   pendientes,       'Esperando procesamiento',            '⏳' if pendientes > 0 else '✅'),
        ('Pedidos En Camino',    enviados,         'En tránsito hacia cliente',          '🚚' if enviados > 0 else '—'),
        ('Pedidos Entregados',   entregados,       'Completados exitosamente',           '✅'),
        ('Pedidos Cancelados',   cancelados,       '',                                   '❌' if cancelados > 0 else '✅'),
        ('Total Ventas ($)',     f'${total_ventas:,.0f}', 'Sin pedidos cancelados',     '💰'),
        ('Productos en Stock',   total_productos,  f'{bajo_stock} con stock bajo',       '⚠️' if bajo_stock > 0 else '✅'),
        ('Bodegas Activas',      total_bodegas,    '',                                   '🏭'),
        ('Tiendas Activas',      total_tiendas,    '',                                   '🏪'),
        ('Envíos En Ruta',       envios_en_ruta,   'Seguimiento activo',                 '🚚' if envios_en_ruta > 0 else '—'),
    ]

    for i, (metrica, valor, detalle, estado) in enumerate(metricas, 5):
        _apply_data_row(ws, [metrica, valor, detalle, estado], row=i, alt=i % 2 == 0)

    _set_col_widths(ws, [30, 20, 35, 10])


def _hoja_pedidos(wb, pedidos):
    ws = wb.create_sheet('📦 Pedidos')
    _titulo(ws, 'Detalle de Pedidos', 9)
    _subtitulo(ws, f'Total: {len(pedidos)} pedidos', 9)

    headers = ['ID', 'Cliente', 'Email', 'Teléfono', 'Dirección Entrega', 'Estado', 'Total ($)', 'Tienda', 'Fecha']
    _apply_header_row(ws, headers, row=3)

    ESTADO_ES = {
        'PENDIENTE': 'Pendiente', 'PROCESANDO': 'Procesando',
        'ENVIADO': 'En camino', 'ENTREGADO': 'Entregado', 'CANCELADO': 'Cancelado',
    }

    for i, p in enumerate(pedidos, 4):
        fecha = str(p.get('fecha_creacion', ''))[:10]
        total = float(p.get('total', 0))
        estado = ESTADO_ES.get(p.get('estado', ''), p.get('estado', ''))
        _apply_data_row(ws, [
            p.get('id', ''),
            p.get('cliente', ''),
            p.get('email_cliente', ''),
            p.get('telefono_cliente', ''),
            p.get('direccion_entrega', ''),
            estado,
            f'${total:,.0f}',
            p.get('tienda_nombre', ''),
            fecha,
        ], row=i, alt=i % 2 == 0)

    # Fila de total
    last = len(pedidos) + 4
    ws.cell(row=last, column=6, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=last, column=7, value=f'=SUM(G4:G{last-1})').font = Font(name='Arial', bold=True, size=10)

    _set_col_widths(ws, [6, 22, 28, 14, 35, 14, 14, 20, 12])


def _hoja_inventario(wb, productos):
    ws = wb.create_sheet('📋 Inventario')
    _titulo(ws, 'Inventario de Productos', 8)
    _subtitulo(ws, f'Total: {len(productos)} productos', 8)

    headers = ['ID', 'Nombre', 'Tipo', 'Precio ($)', 'Stock', 'Stock Mínimo', 'Estado Stock', 'Bodega']
    _apply_header_row(ws, headers, row=3)

    for i, p in enumerate(productos, 4):
        stock     = p.get('stock', 0)
        stock_min = p.get('stock_minimo', 5)
        estado    = '⚠️ Bajo' if stock <= stock_min else '✅ OK'
        precio    = float(p.get('precio', 0))
        _apply_data_row(ws, [
            p.get('id', ''),
            p.get('nombre', ''),
            p.get('tipo', ''),
            f'${precio:,.0f}',
            stock,
            stock_min,
            estado,
            p.get('bodega_nombre', ''),
        ], row=i, alt=i % 2 == 0)

    _set_col_widths(ws, [6, 30, 12, 14, 10, 14, 14, 20])


def _hoja_bodegas_tiendas(wb, bodegas, tiendas):
    ws = wb.create_sheet('🏭 Bodegas y Tiendas')

    # ── Bodegas ──
    _titulo(ws, 'Bodegas y Tiendas', 5)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.cell(row=2, column=1, value=f'Bodegas: {len(bodegas)} | Tiendas: {len(tiendas)}').font = Font(name='Arial', size=9, color='6B7280')

    ws.cell(row=4, column=1, value='BODEGAS').font = Font(name='Arial', bold=True, size=11, color=VERDE)
    _apply_header_row(ws, ['ID', 'Nombre', 'Dirección', 'Capacidad', 'Productos'], row=5, fill_color=VERDE)

    for i, b in enumerate(bodegas, 6):
        _apply_data_row(ws, [
            b.get('id', ''),
            b.get('nombre', ''),
            b.get('direccion', ''),
            b.get('capacidad', 0),
            b.get('total_productos', 0),
        ], row=i, alt=i % 2 == 0)

    sep = len(bodegas) + 8
    ws.cell(row=sep, column=1, value='TIENDAS').font = Font(name='Arial', bold=True, size=11, color=VERDE)
    _apply_header_row(ws, ['ID', 'Nombre', 'Dirección', 'Ciudad', 'Pedidos'], row=sep + 1, fill_color=VERDE)

    for i, t in enumerate(tiendas, sep + 2):
        _apply_data_row(ws, [
            t.get('id', ''),
            t.get('nombre', ''),
            t.get('direccion', ''),
            t.get('ciudad', ''),
            t.get('total_pedidos', 0),
        ], row=i, alt=i % 2 == 0)

    _set_col_widths(ws, [6, 28, 35, 16, 12])


def _hoja_envios(wb, envios):
    ws = wb.create_sheet('🚚 Envíos')
    _titulo(ws, 'Estado de Envíos', 8)
    _subtitulo(ws, f'Total: {len(envios)} envíos', 8)

    headers = ['ID', 'Pedido', 'Tipo', 'Estado', 'Destino', 'Distancia (km)', 'Duración (min)', 'Conductor']
    _apply_header_row(ws, headers, row=3)

    ESTADO_ES = {
        'PENDIENTE': 'Pendiente', 'EN_RUTA': 'En ruta',
        'COMPLETADO': 'Completado', 'FALLIDO': 'Fallido', 'CANCELADO': 'Cancelado',
    }
    TIPO_ES = {'ESTANDAR': 'Estándar', 'EXPRESS': 'Express', 'PROGRAMADO': 'Programado'}

    for i, e in enumerate(envios, 4):
        estado = ESTADO_ES.get(e.get('estado', ''), e.get('estado', ''))
        tipo   = TIPO_ES.get(e.get('tipo', ''), e.get('tipo', ''))
        dist   = e.get('distancia_km', '')
        dur    = e.get('duracion_min', '')
        _apply_data_row(ws, [
            e.get('id', ''),
            e.get('pedido_id', ''),
            tipo,
            estado,
            e.get('destino_nombre', ''),
            f'{float(dist):.1f}' if dist else '',
            dur or '',
            e.get('conductor_nombre', ''),
        ], row=i, alt=i % 2 == 0)

    _set_col_widths(ws, [6, 10, 12, 14, 35, 16, 16, 22])


# ─── Vista ────────────────────────────────────────────────────────────────────

class ExportarExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not OPENPYXL_OK:
            return HttpResponse(
                'openpyxl no está instalado. Ejecuta: pip install openpyxl',
                status=500
            )

        try:
            empresa_rut = ''
            try:
                empresa_rut = request.auth.get('empresa_rut', '') if request.auth else ''
            except Exception:
                pass

            # Obtener datos de todos los microservicios
            pedidos   = MicroserviceGateway.get_pedidos(empresa_rut=empresa_rut)   or []
            productos = MicroserviceGateway.get_inventario(empresa_rut=empresa_rut) or []
            bodegas   = MicroserviceGateway.get_bodegas(empresa_rut=empresa_rut)    or []
            tiendas   = MicroserviceGateway.get_tiendas(empresa_rut=empresa_rut)    or []
            envios    = MicroserviceGateway.get_envios(empresa_rut=empresa_rut)     or []

        except Exception as exc:
            logger.error('Error obteniendo datos para Excel: %s', exc)
            return HttpResponse(f'Error al obtener datos: {exc}', status=503)

        # Generar Excel
        wb = Workbook()
        _hoja_resumen(wb, pedidos, productos, bodegas, tiendas, envios, empresa_rut)
        _hoja_pedidos(wb, pedidos)
        _hoja_inventario(wb, productos)
        _hoja_bodegas_tiendas(wb, bodegas, tiendas)
        _hoja_envios(wb, envios)

        # Guardar en memoria y retornar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fecha     = datetime.now().strftime('%Y%m%d_%H%M')
        filename  = f'SmartLogix_Reporte_{fecha}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response